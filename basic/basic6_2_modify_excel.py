import openpyxl

# 1) 엑셀 불러오기
fpath = "/Users/naegwonhwang/Documents/오징어게임.xlsx"
wb = openpyxl.load_workbook(fpath)

# 2) 엑셀 세트 선택
ws = wb['참가자']

# 3) 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '성기훈'

# 4) 엑셀 저장하기
wb.save(fpath)